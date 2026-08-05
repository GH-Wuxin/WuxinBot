"""Generate a human-review worksheet from the V1 manual-review sample.

``reports/manual-review-v1.jsonl`` is the canonical 300-window sample; this
script only reshapes it for convenient human scoring (one row per window +
acceptance-criteria columns). It never changes the sample itself.

Two artifacts are produced:
- CSV (UTF-8 BOM): stable machine-readable scoring sheet;
- XLSX: human-friendly sheet with frozen header, wrapped text, autofilter
  and Chinese column names.
"""

from __future__ import annotations

import argparse
import csv
import datetime
import json
import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ACCEPTANCE_COLUMNS = [
    "understandable",         # 1 = 脱离原群聊仍能大致理解，0 = 否（目标 >=80%）
    "effective_interaction",  # 1 = 有效互动，0 = 否（目标 >=75%）
    "trigger_reply_correct",  # 1 = 触发/回复关系正确，0 = 否（目标 >=95%）
    "bot_system_spam_only",   # 1 = 纯 bot/系统/刷屏，0 = 否（目标 <=5%）
    "privacy_leak",           # 1 = 高风险隐私泄露，0 = 否（目标 0）
    "notes",                  # 自由备注
]

CSV_COLUMNS = [
    "window_id",
    "start_utc",
    "dataset",
    "window_type",
    "text_sanitized",
    "annotated_lines",
    "group_id_hash",
    "session_id",
    "split",
    "speaker_count",
    "char_count",
    "osu_keyword_count",
    "has_media",
    "media_dependent",
    "privacy_risk",
    "pii_types",
    "bot_line_count",
    "system_line_count",
    "human_text_line_count",
    "has_bot_output",
    "human_only",
    *ACCEPTANCE_COLUMNS,
]

XLSX_HEADERS = [
    "窗口ID",
    "开始时间(UTC)",
    "数据集",
    "窗口类型",
    "脱敏文本",
    "逐行标注",
    "群哈希",
    "会话ID",
    "分区",
    "发言人数",
    "字符数",
    "osu关键词数",
    "含媒体",
    "媒体依赖",
    "隐私风险",
    "PII类型",
    "Bot行数",
    "系统行数",
    "人类文本行数",
    "含Bot输出",
    "纯人类窗口",
    "可独立理解(1/0)",
    "有效互动(1/0)",
    "触发/回复正确(1/0)",
    "纯Bot/系统/刷屏(1/0)",
    "隐私泄露(1/0)",
    "备注",
]

XLSX_WIDTHS = [
    14,
    20,
    16,
    18,
    80,
    80,
    14,
    22,
    18,
    10,
    10,
    12,
    8,
    10,
    10,
    18,
    10,
    10,
    14,
    10,
    12,
    16,
    16,
    18,
    18,
    12,
    30,
]


def _utc(ts_ms: int) -> str:
    return datetime.datetime.fromtimestamp(ts_ms / 1000, tz=datetime.timezone.utc).strftime(
        "%Y-%m-%d %H:%M:%S"
    )


def _load_rows(
    review_path: pathlib.Path,
    annotated_path: pathlib.Path | None = None,
) -> tuple[list[dict], dict[str, str], dict[str, list[dict]]]:
    with review_path.open("r", encoding="utf-8") as f:
        rows = [json.loads(line) for line in f]
    rows.sort(key=lambda r: (r["group_id_hash"], r["start_timestamp"], r["window_id"]))
    annotated: dict[str, str] = {}
    annotated_lines: dict[str, list[dict]] = {}
    if annotated_path is not None and annotated_path.exists():
        with annotated_path.open("r", encoding="utf-8") as f:
            for line in f:
                rec = json.loads(line)
                annotated[rec["window_id"]] = rec["annotated_text"]
                annotated_lines[rec["window_id"]] = rec.get("annotated_lines", [])
    return rows, annotated, annotated_lines


def _line_counts(
    window_id: str,
    annotated_lines: dict[str, list[dict]],
) -> tuple[int, int, int]:
    lines = annotated_lines.get(window_id, [])
    bot = sum(1 for l in lines if l.get("role") == "bot")
    system = sum(1 for l in lines if l.get("role") == "system")
    human_text = sum(
        1 for l in lines if l.get("role") == "human" and (l.get("text") or "").strip()
    )
    return bot, system, human_text


def _row_values(
    r: dict,
    annotated: dict[str, str],
    annotated_lines: dict[str, list[dict]],
) -> list:
    bot, system, human_text = _line_counts(r["window_id"], annotated_lines)
    return [
        r["window_id"],
        _utc(r["start_timestamp"]),
        r.get("dataset", ""),
        r["window_type"],
        r["text_sanitized"],
        annotated.get(r["window_id"], ""),
        r["group_id_hash"],
        r["session_id"],
        r["split"],
        r["speaker_count"],
        r["char_count"],
        r["osu_keyword_count"],
        "1" if r["has_media"] else "0",
        "1" if r["media_dependent"] else "0",
        r["privacy_risk"],
        "|".join(r["pii_types"]),
        bot,
        system,
        human_text,
        "1" if bot > 0 else "0",
        "1" if bot == 0 and system == 0 else "0",
        *([""] * len(ACCEPTANCE_COLUMNS)),
    ]


def build_sheet(
    review_path: pathlib.Path,
    out_path: pathlib.Path,
    annotated_path: pathlib.Path | None = None,
) -> int:
    """Read the JSONL sample and write a scoring CSV. Returns row count."""
    rows, annotated, annotated_lines = _load_rows(review_path, annotated_path)

    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(CSV_COLUMNS)
        for r in rows:
            writer.writerow(_row_values(r, annotated, annotated_lines))
    return len(rows)


def build_sheet_xlsx(
    review_path: pathlib.Path,
    out_path: pathlib.Path,
    annotated_path: pathlib.Path | None = None,
) -> int:
    """Write a human-friendly XLSX scoring sheet (frozen header + wrap)."""
    rows, annotated, annotated_lines = _load_rows(review_path, annotated_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "审核表"
    ws.append(XLSX_HEADERS)

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    for r in rows:
        ws.append(_row_values(r, annotated, annotated_lines))

    for i, width in enumerate(XLSX_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in ("E", "F"):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.zoomScale = 100
    wb.save(out_path)
    return len(rows)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="community-corpus-review-sheet")
    parser.add_argument(
        "--input",
        type=pathlib.Path,
        default=pathlib.Path("reports/manual-review-v1.jsonl"),
    )
    parser.add_argument(
        "--output",
        type=pathlib.Path,
        default=pathlib.Path("reports/manual-review-v1-review-sheet.csv"),
    )
    parser.add_argument(
        "--output-xlsx",
        type=pathlib.Path,
        default=pathlib.Path("reports/manual-review-v1-review-sheet.xlsx"),
    )
    parser.add_argument(
        "--annotated",
        type=pathlib.Path,
        default=pathlib.Path("reports/manual-review-v1-annotated.jsonl"),
        help="optional annotated lines (see review_annotate.py)",
    )
    args = parser.parse_args(argv)
    n = build_sheet(args.input, args.output, args.annotated)
    nx = build_sheet_xlsx(args.input, args.output_xlsx, args.annotated)
    print(f"[review-sheet] {n} rows -> {args.output}")
    print(f"[review-sheet] {nx} rows -> {args.output_xlsx}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
