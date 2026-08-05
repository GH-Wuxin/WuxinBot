"""V1 pipeline tests."""

from __future__ import annotations

import hashlib
import json
import pathlib
import shutil
import tempfile
import unittest

import pyarrow as pa
import pyarrow.parquet as pq

from community_corpus.adapters import clean_text, parse_qce_line
from community_corpus.config import Config
from community_corpus.v1.full_import import find_source_exports, run_full_import
from community_corpus.v1.full_import import BOT_UINS, is_bot_output_like
from community_corpus.v1.reader import load_sender_names
from community_corpus.v1.report_v1 import sample_manual_review, write_manual_review
from community_corpus.v1.review_annotate import _message_index, annotate_window, render_annotated_lines
from community_corpus.v1.review_quickstart import build_quickstart
from community_corpus.v1.review_sheet import build_sheet, build_sheet_xlsx
from community_corpus.v1.sanitize import sanitize_text
from community_corpus.v1.security_fixtures import get_security_fixtures
from community_corpus.v1.sessions import build_sessions
from community_corpus.v1.splits import apply_splits, session_splits
from community_corpus.v1.windows import (
    TIER_STYLE_READY,
    TIER_CONTEXTUAL_STYLE,
    TIER_BOT_INTERACTION,
    TIER_AMBIENT_CHAT,
    TIER_PRIVATE_REJECTED,
    USAGE_TIERS,
    build_windows,
    _classify_tier,
    finalize_window_texts,
    write_windows,
)
from community_corpus.v1.windows import _cluster_overlapping_windows, retrieval_dedupe_windows
from community_corpus.v1.review_precheck import run_precheck, _scan_pii

FIXTURES = pathlib.Path(__file__).parent / "fixtures" / "v1"
SOURCES = [str(FIXTURES / "group_test_200_20260805_000000_chunked_jsonl"), str(FIXTURES / "group_test_201_20260805_000000_chunked_jsonl")]
SALT = "v1-test-salt-2026"


def make_config(tmp: pathlib.Path) -> Config:
    return Config(
        sources=tuple(SOURCES),
        sample_size=50_000,
        seed=20260805,
        salt=SALT,
        output_dir=tmp,
    )


def run_v1(tmp: pathlib.Path, salt: str = SALT):
    cfg = Config(
        sources=tuple(SOURCES),
        sample_size=50_000,
        seed=20260805,
        salt=salt,
        output_dir=tmp,
    )
    full_result = run_full_import(cfg)
    messages = pq.read_table(cfg.normalized_dir / "full" / "messages.parquet").to_pylist()
    sessions = build_sessions(messages, gap_minutes=8)
    table = pq.read_table(cfg.normalized_dir / "full" / "messages.parquet")
    windows, filter_stats = build_windows(table, sessions)
    export_dirs = find_source_exports(cfg.sources)
    refs = set()
    for w in windows:
        for m in w["_messages"]:
            refs.add((m["source_export"], m["source_file"], m["source_offset_bytes"]))
    names = load_sender_names(export_dirs, refs)
    finalize_window_texts(windows, names)
    splits = session_splits(sessions, 20260805)
    apply_splits(windows, splits)
    return cfg, full_result, sessions, windows, filter_stats


class V1PipelineTest(unittest.TestCase):
    def setUp(self):
        self.tmp = pathlib.Path(tempfile.mkdtemp(prefix="community-corpus-v1-test-"))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_full_import_does_not_touch_v0(self):
        cfg = make_config(self.tmp)
        # simulate V0 50k sample parquet
        v0 = self.tmp / "normalized" / "messages.parquet"
        v0.parent.mkdir(parents=True, exist_ok=True)
        table = pa.table({"marker": pa.array(["v0"], type=pa.string())})
        pq.write_table(table, v0)
        v0_hash_before = hashlib.sha256(v0.read_bytes()).hexdigest()

        full_result = run_full_import(cfg)
        self.assertEqual(full_result["totalMessages"], 35)  # 31 + 4
        v0_hash_after = hashlib.sha256(v0.read_bytes()).hexdigest()
        self.assertEqual(v0_hash_before, v0_hash_after)
        self.assertTrue((self.tmp / "normalized" / "full" / "messages.parquet").exists())

    def test_full_import_report_written(self):
        cfg = make_config(self.tmp)
        result = run_full_import(cfg)
        p = cfg.reports_dir / "full-import-report.json"
        self.assertTrue(p.exists())
        data = json.loads(p.read_text(encoding="utf-8"))
        self.assertEqual(data["summary"]["totalMessages"], result["totalMessages"])
        self.assertEqual(data["summary"]["uniqueMessages"], 35)
        for key in ("typeCounts", "piiTypeCounts", "perSource", "parseFailures"):
            self.assertIn(key, data)

    def test_salt_stable(self):
        cfg, full_result, sessions, windows, _ = run_v1(self.tmp)
        rows = pq.read_table(cfg.normalized_dir / "full" / "messages.parquet").to_pylist()
        first = rows[0]["sender_id_hash"]
        # rerun with same salt -> same hash
        cfg2, _, _, _, _ = run_v1(self.tmp)
        rows2 = pq.read_table(cfg2.normalized_dir / "full" / "messages.parquet").to_pylist()
        self.assertEqual(first, rows2[0]["sender_id_hash"])
        # different salt -> different hash
        cfg3, _, _, _, _ = run_v1(self.tmp, salt="other-salt")
        rows3 = pq.read_table(cfg3.normalized_dir / "full" / "messages.parquet").to_pylist()
        self.assertNotEqual(first, rows3[0]["sender_id_hash"])

    def test_sessions_do_not_cross_group(self):
        _, _, sessions, _, _ = run_v1(self.tmp)
        for s in sessions:
            self.assertTrue(s["group_id_hash"])
        # every session's messages belong to one group hash (checked via construction)
        group_of_session = {s["session_id"]: s["group_id_hash"] for s in sessions}
        self.assertEqual(len(group_of_session), len(sessions))

    def test_time_segmentation_and_deterministic(self):
        cfg, _, sessions, _, _ = run_v1(self.tmp)
        # gap > 8min between 2021 (t20) and 2022 (t31) => 2 sessions in group test_200
        g200 = [s for s in sessions if s["group_id_hash"]]
        all_ids = [s["session_id"] for s in sessions]
        # messages 2021 and 2022 must be in different sessions
        s2021 = next(s for s in sessions if "2021" in s["message_ids"])
        s2022 = next(s for s in sessions if "2022" in s["message_ids"])
        self.assertNotEqual(s2021["session_id"], s2022["session_id"])
        # rerun gives identical sessions
        _, _, sessions2, _, _ = run_v1(self.tmp)
        self.assertEqual(
            [(s["session_id"], s["message_ids"]) for s in sessions],
            [(s["session_id"], s["message_ids"]) for s in sessions2],
        )

    def test_reply_chain_order_and_depth(self):
        _, _, sessions, windows, _ = run_v1(self.tmp)
        reply_windows = [w for w in windows if w["window_type"] == "reply_chain"]
        # 2002 replies to 2001
        w2002 = next((w for w in reply_windows if w["trigger_message_id"] == "2002"), None)
        self.assertIsNotNone(w2002)
        ids = w2002["message_ids"]
        self.assertEqual(ids[0], "2001")
        self.assertEqual(ids[1], "2002")
        self.assertLessEqual(w2002["reply_depth"], 4)
        # 3003 replies to 3002
        w3003 = next((w for w in reply_windows if w["trigger_message_id"] == "3003"), None)
        self.assertIsNotNone(w3003)
        self.assertEqual(w3003["message_ids"][0], "3002")
        self.assertEqual(w3003["message_ids"][1], "3003")

    def test_deep_reply_chain_truncated(self):
        # build synthetic rows for a 6-level chain and verify depth cap
        rows = []
        base = 1785400000000
        for i in range(7):
            mid = f"c{i}"
            reply = f"c{i-1}" if i > 0 else None
            rows.append(
                {
                    "message_id": mid,
                    "group_id_hash": "g1",
                    "sender_id_hash": f"s{i}",
                    "timestamp": base + i * 60_000,
                    "seq": str(i),
                    "reply_to_id": reply,
                    "message_type": "text",
                    "text_raw": f"msg {i}",
                    "text_clean": f"msg {i}",
                    "has_media": False,
                    "media_type": "none",
                    "is_bot": False,
                    "is_system": False,
                    "bot_output_like": False,
                    "source_file": "chunks/chunk_0001.jsonl",
                    "source_offset": i + 1,
                    "source_offset_bytes": 0,
                    "source_export": "group_x",
                }
            )
        table = pa.table(
            {
                "message_id": pa.array([r["message_id"] for r in rows], type=pa.string()),
                "group_id_hash": pa.array([r["group_id_hash"] for r in rows], type=pa.string()),
                "sender_id_hash": pa.array([r["sender_id_hash"] for r in rows], type=pa.string()),
                "timestamp": pa.array([r["timestamp"] for r in rows], type=pa.int64()),
                "seq": pa.array([r["seq"] for r in rows], type=pa.string()),
                "reply_to_id": pa.array([r["reply_to_id"] for r in rows], type=pa.string()),
                "message_type": pa.array([r["message_type"] for r in rows], type=pa.string()),
                "text_raw": pa.array([r["text_raw"] for r in rows], type=pa.string()),
                "text_clean": pa.array([r["text_clean"] for r in rows], type=pa.string()),
                "has_media": pa.array([r["has_media"] for r in rows], type=pa.bool_()),
                "media_type": pa.array([r["media_type"] for r in rows], type=pa.string()),
                "is_bot": pa.array([r["is_bot"] for r in rows], type=pa.bool_()),
                "is_system": pa.array([r["is_system"] for r in rows], type=pa.bool_()),
                "bot_output_like": pa.array([r["bot_output_like"] for r in rows], type=pa.bool_()),
                "source_file": pa.array([r["source_file"] for r in rows], type=pa.string()),
                "source_offset": pa.array([r["source_offset"] for r in rows], type=pa.int64()),
                "source_offset_bytes": pa.array([r["source_offset_bytes"] for r in rows], type=pa.int64()),
                "source_export": pa.array([r["source_export"] for r in rows], type=pa.string()),
            }
        )
        sessions = [
            {
                "session_id": "g1-s000001",
                "group_id_hash": "g1",
                "start_timestamp": base,
                "end_timestamp": base + 6 * 60_000,
                "message_ids": [f"c{i}" for i in range(7)],
                "context_message_ids": [],
                "message_count": 7,
                "speaker_count": 7,
            }
        ]
        windows, _ = build_windows(table, sessions)
        chains = [w for w in windows if w["window_type"] == "reply_chain"]
        w6 = next(w for w in chains if w["trigger_message_id"] == "c6")
        # ancestors: c5,c4,c3,c2 (4 levels), then c6; c1 excluded
        self.assertEqual(w6["reply_depth"], 4)
        self.assertIn("c2", w6["message_ids"])
        self.assertIn("c5", w6["message_ids"])
        self.assertNotIn("c1", w6["message_ids"])
        self.assertEqual(w6["message_ids"][0], "c2")
        self.assertEqual(w6["message_ids"][-1], "c6")

    def test_windows_traceable_to_raw(self):
        _, _, _, windows, _ = run_v1(self.tmp)
        export_by_name = {p.name: p for p in find_source_exports(SOURCES)}
        checked = 0
        for w in windows:
            for ref in w["source_refs"]:
                p = export_by_name[ref["source_export"]] / ref["source_file"]
                with p.open("r", encoding="utf-8", newline="") as f:
                    f.seek(ref["source_offset_bytes"])
                    line = f.readline()
                self.assertIn(ref["message_id"], line)
                checked += 1
        self.assertGreater(checked, 0)

    def test_text_sanitized_has_no_known_pii(self):
        _, _, _, windows, _ = run_v1(self.tmp)
        secrets = [
            "123456789",
            "13800138000",
            "a@b.com",
            "192.168.1.1",
            "sk-abcdef1234567890",
            "110101199003077777",
            "北京大学",
            "jq.qq.com",
            "甲选手",
            "乙选手",
        ]
        for w in windows:
            for s in secrets:
                self.assertNotIn(s, w["text_sanitized"], f"{s} leaked in {w['window_id']}")

    def test_osu_numbers_preserved(self):
        _, _, _, windows, _ = run_v1(self.tmp)
        joined = "\n".join(w["text_sanitized"] for w in windows)
        self.assertIn("pp 12345", joined)
        self.assertIn("rank 34567", joined)
        self.assertIn("acc 99.1%", joined)
        self.assertIn("beatmap 3456789", joined)
        self.assertIn("bp 654321", joined)

    def test_same_session_same_split(self):
        _, _, _, windows, _ = run_v1(self.tmp)
        split_by_session: dict[str, set[str]] = {}
        for w in windows:
            split_by_session.setdefault(w["session_id"], set()).add(w["split"])
        for sid, splits in split_by_session.items():
            self.assertEqual(len(splits), 1, f"session {sid} spans splits {splits}")

    def test_near_duplicate_windows_clustered(self):
        _, _, _, windows, _ = run_v1(self.tmp)
        for w in windows:
            self.assertTrue(w["overlap_cluster_id"])
            self.assertIn(w["usage_tier"], USAGE_TIERS)
        deduped = retrieval_dedupe_windows(windows)
        by_session: dict[str, list[dict]] = {}
        for w in deduped:
            by_session.setdefault(w["session_id"], []).append(w)
        pairs = 0
        for session_windows in by_session.values():
            ids = [set(w["message_ids"]) for w in session_windows]
            for i in range(len(ids)):
                for j in range(i + 1, len(ids)):
                    union = len(ids[i] | ids[j])
                    if union and len(ids[i] & ids[j]) / union >= 0.8:
                        pairs += 1
        self.assertEqual(pairs, 0)

    def test_cluster_representative_is_longest(self):
        def w(i: int, mids: list[str], start: int) -> dict:
            return {
                "window_id": f"W{i:08d}",
                "session_id": "g1-s1",
                "start_timestamp": start,
                "end_timestamp": start + 10,
                "trigger_message_id": "t1",
                "message_ids": mids,
                "_messages": [{"text_clean": ""}],
            }

        windows = [
            w(0, ["a", "b", "c", "d"], 100),
            w(1, ["a", "b", "c"], 100),
            w(2, ["x", "y", "z"], 200),
        ]
        clustered, stats = _cluster_overlapping_windows(windows)
        self.assertEqual(len(clustered), 3)  # nothing is deleted
        self.assertEqual(stats["cluster_count"], 1)
        self.assertEqual(stats["clustered_window_count"], 2)
        self.assertEqual(stats["max_cluster_size"], 2)
        self.assertEqual(
            clustered[0]["overlap_cluster_id"],
            clustered[1]["overlap_cluster_id"],
        )
        self.assertTrue(clustered[0]["overlap_cluster_representative"])
        self.assertFalse(clustered[1]["overlap_cluster_representative"])
        self.assertNotEqual(
            clustered[0]["overlap_cluster_id"],
            clustered[2]["overlap_cluster_id"],
        )
        kept = retrieval_dedupe_windows(clustered)
        self.assertEqual([x["window_id"] for x in kept], ["W00000000", "W00000002"])
        self.assertEqual(sum(stats["reasons"].values()), 1)

    def test_deterministic_repeat(self):
        cfg, _, _, windows, _ = run_v1(self.tmp)
        first = [
            (w["window_id"], w["window_type"], w["message_ids"], w["text_sanitized"], w["split"])
            for w in windows
        ]
        cfg2, _, _, windows2, _ = run_v1(self.tmp)
        second = [
            (w["window_id"], w["window_type"], w["message_ids"], w["text_sanitized"], w["split"])
            for w in windows2
        ]
        self.assertEqual(first, second)

    def test_sanitize_unit(self):
        text, types, confidence, risk = sanitize_text("我 QQ 是 123456789", {})
        self.assertNotIn("123456789", text)
        self.assertIn("qq", types)
        text, types, confidence, risk = sanitize_text("beatmap 3456789 pp 12345", {})
        self.assertIn("beatmap 3456789 pp 12345", text)
        self.assertEqual(types, [])

    def test_sanitize_credential_url_and_token(self):
        url = (
            "https://osugaming.sekai.team/login?token="
            "WDE6hJQ28DEsv6cHRkRVMcSRdXnsQPp%2FWNLtshr4ebFEd9YU5HxRS%2BSX8rB4CqiwSfDrMO%2BAAI"
        )
        text, types, _, _ = sanitize_text(url, {})
        self.assertNotIn("token=", text)
        self.assertNotIn("%2F", text)
        self.assertIn("credential", types)

    def test_sanitize_group_number_discord_forward_invite(self):
        text, types, _, _ = sanitize_text("osu！新人FPS群722050824", {})
        self.assertNotIn("722050824", text)
        self.assertIn("qq", types)

        profile = "Occupation: 狂妄之人\nDiscord: bget3066\nWebsite: https://live.bilibili.com/30286637"
        text, types, _, _ = sanitize_text(profile, {})
        self.assertNotIn("bget3066", text)
        self.assertNotIn("live.bilibili.com", text)
        self.assertNotIn("狂妄之人", text)
        self.assertIn("profile", types)

        fwd = (
            "[转发消息: 13条]\n  Toriesta: 宝宝你好烧\n"
            "🌕🌕🌕🌕🌕🌕🌕🌕🌕🌕\n"
            "  佐佐佑佑: 宝宝你好烧\n"
            "  PC: 44,602"
        )
        text, types, _, _ = sanitize_text(fwd, {})
        self.assertNotIn("Toriesta", text)
        self.assertNotIn("佐佐佑佑", text)
        self.assertIn("<NICK>", text)
        self.assertIn("PC: 44,602", text)
        self.assertIn("nickname", types)

        link = "https://endfield.hypergryph.com/activity/x?invite_code=Z029FJ0QD59RRMM6&share_type=link"
        text, types, _, _ = sanitize_text(link, {})
        self.assertNotIn("Z029FJ0QD59RRMM6", text)
        self.assertIn("<INVITE>", text)
        self.assertIn("invite", types)

        for invite_link in (
            "https://oopz.cn/i/jUg00i",
            "https://discord.gg/AbCdEfG",
            "https://kook.app/invite/AbCdEfG",
            "https://kookapp.cn/invite/AbCdEfG",
        ):
            text, types, _, _ = sanitize_text(invite_link, {})
            self.assertNotIn(invite_link, text)
            self.assertIn("<INVITE>", text)
            self.assertIn("invite", types)

    def test_sanitize_bare_qq_number_keeps_public_url_numbers(self):
        for raw, secret in (
            ("S2 712531032", "712531032"),
            ("群的话是这个 1039477458，进来别骂我菜", "1039477458"),
            ("绑定到 948117351 上！", "948117351"),
            ("!ymra 121189283", "121189283"),
        ):
            text, types, _, _ = sanitize_text(raw, {})
            self.assertNotIn(secret, text)
            self.assertIn("<QQ_NUMBER>", text)
            self.assertIn("qq", types)

        # public web/score ids embedded in URL paths must survive
        for raw, keep in (
            ("https://osu.ppy.sh/scores/6082266269", "6082266269"),
            ("https://live.bilibili.com/1714609201", "1714609201"),
            ("https://www.pixiv.net/i/142389334", "142389334"),
        ):
            text, types, _, _ = sanitize_text(raw, {})
            self.assertIn(keep, text)
            self.assertNotIn("qq", types)

        # currency-style numbers with units/decimal parts must survive
        for raw, keep in (
            ("汇率2147483648usd", "2147483648usd"),
            ("CNY 470548.48", "470548.48"),
        ):
            text, types, _, _ = sanitize_text(raw, {})
            self.assertIn(keep, text)
            self.assertNotIn("qq", types)

        # inviterUid is a private URL param and must be redacted
        text, types, _, _ = sanitize_text(
            "https://example.com/invite?inviterUid=275180748&from=group", {}
        )
        self.assertNotIn("275180748", text)
        self.assertIn("<INVITE>", text)
        self.assertIn("invite", types)

    def test_sanitize_empty_profile_field_and_bracket_mention(self):
        text, types, _, _ = sanitize_text("Interests:\nS1 怎么又掉出5w了", {})
        self.assertIn("Interests:", text)
        self.assertNotIn("profile", types)

        text, types, _, _ = sanitize_text(
            "[内联键盘][Markdown消息]@[语音通话] 语音通话已结束",
            {"语音通话": "S2"},
        )
        self.assertNotIn("@[", text)
        self.assertIn("<MENTION>", text)
        self.assertIn("mention", types)

    def test_reply_sender_name_stripped(self):
        line = (
            '{"id":"r1","seq":"1","timestamp":1785400000000,'
            '"sender":{"uid":"u1","uin":"1111111111","name":"xianfishyu","groupCard":"xianfishyu"},'
            '"type":"reply","content":{"text":"[回复消息]@-inui sana-｜吉祥物 我越看越觉得ppy像是那个哥布林",'
            '"html":"","elements":[{"type":"reply","data":{"messageId":"m1","referencedMessageId":"m1",'
            '"senderUin":"630060047","senderName":"-inui sana-｜吉祥物","content":"[图片]"}}],'
            '"resources":[],"mentions":[]},"recalled":false,"system":false}'
        )
        rec = parse_qce_line(line)
        self.assertIn("-inui sana-｜吉祥物", rec.reply_sender_names)
        cleaned = clean_text(rec.text_raw, list(rec.reply_sender_names))
        self.assertNotIn("sana-｜吉祥物", cleaned)
        self.assertIn("我越看越觉得ppy像是那个哥布林", cleaned)

    def test_lone_surrogate_repaired(self):
        from community_corpus.adapters import parse_qce_line

        line = (
            '{"id":"9001","seq":"9001","timestamp":1785400000000,'
            '"sender":{"uid":"u_a","uin":"1111111111","name":"\\ud83c 测试","groupCard":"\\ud83c 测试"},'
            '"type":"text","content":{"text":"坏字符 \\ud83c 结尾","html":"","elements":[],"resources":[],"mentions":[]},'
            '"recalled":false,"system":false}'
        )
        rec = parse_qce_line(line)
        self.assertNotIn("\ud83c", rec.text_raw)
        self.assertNotIn("\ud83c", rec.sender_name)
        self.assertIn("\ufffd", rec.text_raw)

    def test_manual_review_includes_all_high_risk(self):
        windows = []
        for i in range(25):
            risk = "high" if i < 5 else "low"
            windows.append(
                {
                    "window_id": f"W{i:08d}",
                    "window_type": "temporal_burst",
                    "usage_tier": "style_ready",
                    "overlap_cluster_id": f"W{i:08d}",
                    "group_id_hash": "g1",
                    "session_id": "g1-s000001",
                    "start_timestamp": 1785400000000 + i,
                    "human_message_count": 2,
                    "privacy_risk": risk,
                }
            )
        sample = sample_manual_review(windows, seed=20260805, n=10)
        self.assertEqual(sum(1 for w in sample if w["privacy_risk"] == "high"), 5)
        self.assertEqual(len(sample), 10)

    def test_review_sheet_written(self):
        src = self.tmp / "sample.jsonl"
        src.write_text(
            json.dumps(
                {
                    "window_id": "W00000001",
                    "window_type": "temporal_burst",
                    "usage_tier": "style_ready",
                    "overlap_cluster_id": "W00000001",
                    "group_id_hash": "g1",
                    "session_id": "g1-s000001",
                    "split": "train_candidate",
                    "start_timestamp": 1785400000000,
                    "speaker_count": 2,
                    "char_count": 20,
                    "osu_keyword_count": 0,
                    "has_media": False,
                    "media_dependent": False,
                    "privacy_risk": "low",
                    "pii_types": [],
                    "text_sanitized": "S1 hi\nS2 hello\x14",
                },
                ensure_ascii=False,
            )
            + "\n",
            encoding="utf-8",
        )
        out = self.tmp / "sheet.csv"
        n = build_sheet(src, out)
        self.assertEqual(n, 1)
        text = out.read_text(encoding="utf-8-sig")
        self.assertIn("window_id", text)
        self.assertIn("usage_tier", text)
        self.assertIn("overlap_cluster_id", text)
        self.assertIn("understandable", text)
        self.assertIn("has_bot_output", text)
        self.assertIn("human_only", text)
        self.assertIn("S1 hi", text)

    def test_review_annotate_roles(self):
        cfg, _, _, windows, _ = run_v1(self.tmp)
        msg_index = _message_index(cfg.normalized_dir / "full" / "messages.parquet")
        exports_root = FIXTURES / "v1"
        found_bot = False
        for w in windows:
            lines = annotate_window(w, msg_index, exports_root)
            text = render_annotated_lines(lines)
            self.assertIn("S", text)
            self.assertTrue(all(l["speaker_label"] != "?" for l in lines))
            self.assertTrue(all("sender_name" not in l for l in lines))
            if any(l["role"] == "bot" for l in lines):
                found_bot = True
        self.assertTrue(found_bot)  # fixture u_bot is marked is_bot

    def test_review_quickstart_written(self):
        src = self.tmp / "sample.jsonl"
        ann = self.tmp / "annotated.jsonl"
        out = self.tmp / "quickstart.md"
        src.write_text(
            json.dumps(
                {
                    "window_id": "W00000001",
                    "window_type": "temporal_burst",
                    "group_id_hash": "g1",
                    "session_id": "g1-s000001",
                    "split": "train_candidate",
                    "start_timestamp": 1785400000000,
                    "privacy_risk": "high",
                    "media_dependent": True,
                },
                ensure_ascii=False,
            )
            + "\n",
            encoding="utf-8",
        )
        ann.write_text(
            json.dumps(
                {
                    "window_id": "W00000001",
                    "annotated_lines": [
                        {
                            "role": "human",
                            "text": "！pr",
                            "bot_output_like": False,
                        }
                    ],
                },
                ensure_ascii=False,
            )
            + "\n",
            encoding="utf-8",
        )
        n = build_quickstart(src, ann, out)
        self.assertEqual(n, 1)
        text = out.read_text(encoding="utf-8")
        self.assertIn("W00000001", text)
        self.assertIn("高风险", text)

    def test_bot_output_like(self):
        for uin in (
            "1708547915",  # 忧郁小猫猫
            "1902931474",  # KQN
            "3311470495",  # 天使果果喵
            "3145729213",  # 雨沐 (临时运行)
            "3889016014",  # 雨沐 (原版)
            "2818054860",  # 小幽幽子
            "3889001246",  # 幽幽子
            "3929371650",  # Nikaidou Shinku
            "1750011571",  # ATRI1024
            "1078589506",  # 全自助火化机
            "814992458",  # 遠野幻想物語
            "1335734629",  # 白菜V2.1 (recent-score card bot)
            "1020640876",  # 白菜V2.1 (binding error bot)
            "2225126759",  # Lazybot测试机
        ):
            self.assertIn(uin, BOT_UINS)
        self.assertTrue(is_bot_output_like("Miku的个人信息—osu!\n\n5025.66pp 表现\n#12345"))
        self.assertTrue(is_bot_output_like("ElicyAnn的个人信息—mania\n\n7832.35pp 表现\n#9197"))
        self.assertTrue(is_bot_output_like("box1n的个人信息—taiko\n\n1358.26pp 表现\n#28871"))
        self.assertTrue(
            is_bot_output_like(
                "Hinanawi Tenshi 与 Firika 的 replay 轨迹\n相似度: 0.88%"
            )
        )
        self.assertTrue(
            is_bot_output_like(
                "S2 与 [SHK]qazoop 的 replay 检测\n相似度: -17.31%"
            )
        )
        self.assertTrue(is_bot_output_like("少女祈祷中..."))
        self.assertTrue(
            is_bot_output_like(
                "主要数据已更新完毕，pp+数据正在后台更新，请稍后使用info功能查看结果。"
            )
        )
        self.assertTrue(is_bot_output_like("pr1mary头像已更新"))
        self.assertTrue(is_bot_output_like("查@某人 结果：Someone的个人信息—osu!\n\n0pp"))
        self.assertTrue(is_bot_output_like("wilson1125的bp类型\nAim:26.01% (23张)"))
        self.assertTrue(is_bot_output_like("正在获取pp+数据，请稍等。。"))
        self.assertTrue(is_bot_output_like("[内联键盘][Markdown消息]@sakura 绑定成功"))
        self.assertTrue(is_bot_output_like("[Lazybot] 没这B人: name=dragon"))
        self.assertTrue(
            is_bot_output_like("根据 BP 关联度，在 osu! 模式给 Wux1n 推荐的图如下：")
        )
        self.assertFalse(is_bot_output_like("看看 bp"))
        self.assertFalse(is_bot_output_like("个人信息查询很好用"))
        cfg, _, _, windows, _ = run_v1(self.tmp)
        rows = pq.read_table(cfg.normalized_dir / "full" / "messages.parquet").to_pylist()
        miku = next(r for r in rows if r["message_id"] == "2031")
        self.assertTrue(miku["bot_output_like"])
        self.assertTrue(
            any(
                w["usage_tier"] == TIER_BOT_INTERACTION
                and any(m["message_id"] == "2031" for m in w["_messages"])
                for w in windows
            )
        )

    def test_window_tiers_assigned(self):
        _, _, _, windows, _ = run_v1(self.tmp)
        tiers = {w["usage_tier"] for w in windows}
        self.assertTrue(tiers <= USAGE_TIERS)

    def test_spam_media_window_rejected(self):
        def row(mid: str, text: str, has_media: bool = False) -> dict:
            return {
                "message_id": mid,
                "sender_id_hash": f"s{mid}",
                "timestamp": 0,
                "seq": mid,
                "reply_to_id": None,
                "message_type": "text",
                "text_raw": text,
                "text_clean": text,
                "has_media": has_media,
                "media_type": "none",
                "is_bot": False,
                "is_system": False,
                "bot_output_like": False,
                "source_file": "chunks/chunk_0001.jsonl",
                "source_offset": 1,
                "source_offset_bytes": 0,
                "source_export": "group_x",
            }

        spam = [
            row("a", "", True),
            row("b", "牛的"),
            row("c", "牛的"),
            row("d", "牛的"),
        ]
        self.assertEqual(_classify_tier(spam, True, [], "low"), TIER_PRIVATE_REJECTED)

    def test_usage_tier_classification(self):
        def row(mid: str, text: str, sender: str, media: bool = False,
                bot: bool = False, system: bool = False) -> dict:
            return {
                "message_id": mid,
                "sender_id_hash": sender,
                "timestamp": 0,
                "seq": mid,
                "reply_to_id": None,
                "message_type": "text",
                "text_raw": text,
                "text_clean": text,
                "has_media": media,
                "media_type": "none",
                "is_bot": bot,
                "is_system": system,
                "bot_output_like": False,
                "source_file": "chunks/chunk_0001.jsonl",
                "source_offset": 1,
                "source_offset_bytes": 0,
                "source_export": "group_x",
            }

        style = [
            row("a", "今天这张图手感真的不错", "s1"),
            row("b", "确实，我acc到98了", "s2"),
        ]
        self.assertEqual(_classify_tier(style, False, [], "low"), TIER_STYLE_READY)

        media = [
            row("a", "", "s1", media=True),
            row("b", "牛的", "s2"),
        ]
        self.assertEqual(_classify_tier(media, True, [], "low"), TIER_CONTEXTUAL_STYLE)

        ambient = [
            row("a", "6", "s1"),
            row("b", "牛的", "s2"),
        ]
        self.assertEqual(_classify_tier(ambient, False, [], "low"), TIER_AMBIENT_CHAT)

        bot_cmd = [
            row("a", "!pr", "s1"),
            row("b", "Miku的个人信息—osu!\n\n5025pp", "bot", bot=True),
        ]
        self.assertEqual(_classify_tier(bot_cmd, False, [], "low"), TIER_BOT_INTERACTION)

        spaced_cmd = [
            row("a", "! rs", "s1"),
            row("b", "官网链接：https://osu.ppy.sh/b/5405851\n模式：Standard", "s2"),
        ]
        self.assertEqual(
            _classify_tier(spaced_cmd, False, [], "low"), TIER_BOT_INTERACTION
        )

        spam = [
            row("a", "牛的", "s1"),
            row("b", "牛的", "s2"),
            row("c", "牛的", "s3"),
        ]
        self.assertEqual(_classify_tier(spam, False, [], "low"), TIER_PRIVATE_REJECTED)

        high_risk = [
            row("a", "今天手感不错", "s1"),
            row("b", "确实", "s2"),
        ]
        self.assertEqual(
            _classify_tier(high_risk, False, ["phone"], "high"),
            TIER_PRIVATE_REJECTED,
        )

        sys_only = [
            row("a", "系统提示", "sys", system=True),
            row("b", "", "sys2", system=True),
        ]
        self.assertEqual(
            _classify_tier(sys_only, False, [], "low"), TIER_PRIVATE_REJECTED
        )

        sys_with_human = [
            row("a", "系统提示", "sys", system=True),
            row("b", "我也觉得", "s1"),
        ]
        self.assertEqual(
            _classify_tier(sys_with_human, False, [], "low"),
            TIER_BOT_INTERACTION,
        )

        bot_with_spam_reactions = [
            row("a", "！pr", "s1"),
            row("b", "牛的", "s2"),
            row("c", "牛的", "s3"),
            row("d", "牛的", "s4"),
        ]
        self.assertEqual(
            _classify_tier(bot_with_spam_reactions, False, [], "low"),
            TIER_PRIVATE_REJECTED,
        )

        forward_block = [
            row("a", "竹姐好忙", "s1"),
            row("b", "[转发消息: 6条]\n  <NICK>: [图片]\n  <NICK>: [图片]", "s2"),
        ]
        self.assertEqual(
            _classify_tier(forward_block, False, [], "low"),
            TIER_CONTEXTUAL_STYLE,
        )

    def test_security_fixtures_all_redacted(self):
        for fx in get_security_fixtures():
            raw = fx["raw"]
            prepare = fx.get("prepare")
            if prepare is not None:
                raw = prepare(raw)
            text, types, confidence, risk = sanitize_text(raw, {})
            for secret in fx["must_not_contain"]:
                self.assertNotIn(secret, text, f"{fx['name']}: {secret!r} leaked")
            self.assertEqual(_scan_pii(text), set(), fx["name"])
            if fx.get("expect_types"):
                self.assertTrue(
                    set(fx["expect_types"]) & set(types),
                    f"{fx['name']}: expected {fx['expect_types']} in {types}",
                )

    def test_precheck_privacy_first_gates(self):
        def window(i: int, text: str = "S1 hi\nS2 hello") -> dict:
            return {
                "window_id": f"W{i:08d}",
                "usage_tier": "style_ready",
                "overlap_cluster_id": f"W{i:08d}",
                "overlap_cluster_representative": True,
                "message_ids": [f"m{i}"],
                "media_dependent": False,
                "human_message_count": 2,
                "bot_message_count": 0,
                "bot_output_count": 0,
                "privacy_risk": "low",
                "text_sanitized": text,
                "source_refs": "[]",
            }

        windows = [window(i) for i in range(300)]
        table = pa.table(
            {
                "window_id": pa.array([w["window_id"] for w in windows], type=pa.string()),
                "usage_tier": pa.array([w["usage_tier"] for w in windows], type=pa.string()),
                "overlap_cluster_id": pa.array(
                    [w["overlap_cluster_id"] for w in windows], type=pa.string()
                ),
                "overlap_cluster_representative": pa.array(
                    [w["overlap_cluster_representative"] for w in windows], type=pa.bool_()
                ),
                "message_ids": pa.array(
                    [w["message_ids"] for w in windows], type=pa.list_(pa.string())
                ),
                "media_dependent": pa.array(
                    [w["media_dependent"] for w in windows], type=pa.bool_()
                ),
                "human_message_count": pa.array(
                    [w["human_message_count"] for w in windows], type=pa.int64()
                ),
                "bot_message_count": pa.array(
                    [w["bot_message_count"] for w in windows], type=pa.int64()
                ),
                "bot_output_count": pa.array(
                    [w["bot_output_count"] for w in windows], type=pa.int64()
                ),
                "privacy_risk": pa.array(
                    [w["privacy_risk"] for w in windows], type=pa.string()
                ),
                "text_sanitized": pa.array(
                    [w["text_sanitized"] for w in windows], type=pa.string()
                ),
                "source_refs": pa.array(
                    [w["source_refs"] for w in windows], type=pa.string()
                ),
            }
        )
        win_path = self.tmp / "windows.parquet"
        pq.write_table(table, win_path)
        review_path = self.tmp / "review.jsonl"
        review_path.write_text(
            "".join(
                json.dumps(w, ensure_ascii=False) + "\n"
                for w in windows
            ),
            encoding="utf-8",
        )

        result = run_precheck(
            review_path, win_path, None, None, full_scan=False
        )
        self.assertTrue(result["passed"])
        self.assertEqual(result["gates"]["privacyLeakCount"], 0)
        self.assertEqual(result["gates"]["sampleClusterDupes"], 0)
        self.assertEqual(result["gates"]["usageTierMissing"], 0)
        self.assertEqual(result["gates"]["overlapClusterMissing"], 0)

        # privacy is the only veto: one leaked phone number fails the gate
        windows[0]["text_sanitized"] = "S1 hi 13800138000"
        review_path.write_text(
            "".join(
                json.dumps(w, ensure_ascii=False) + "\n"
                for w in windows
            ),
            encoding="utf-8",
        )
        result = run_precheck(review_path, win_path, None, None, full_scan=False)
        self.assertFalse(result["passed"])
        self.assertEqual(result["gates"]["privacyLeakCount"], 1)

        # two non-high-risk windows sharing a cluster fail the review gate
        windows[0]["text_sanitized"] = "S1 hi\nS2 hello"
        windows[1]["overlap_cluster_id"] = windows[0]["overlap_cluster_id"]
        table = table.set_column(
            table.schema.get_field_index("overlap_cluster_id"),
            "overlap_cluster_id",
            pa.array([w["overlap_cluster_id"] for w in windows], type=pa.string()),
        )
        pq.write_table(table, win_path)
        review_path.write_text(
            "".join(
                json.dumps(w, ensure_ascii=False) + "\n"
                for w in windows
            ),
            encoding="utf-8",
        )
        result = run_precheck(review_path, win_path, None, None, full_scan=False)
        self.assertFalse(result["passed"])
        self.assertEqual(result["gates"]["sampleClusterDupes"], 1)

    def test_review_sheet_xlsx_written(self):
        src = self.tmp / "sample.jsonl"
        ann = self.tmp / "annotated.jsonl"
        out = self.tmp / "sheet.xlsx"
        src.write_text(
            json.dumps(
                {
                    "window_id": "W00000001",
                    "window_type": "temporal_burst",
                    "usage_tier": "style_ready",
                    "overlap_cluster_id": "W00000001",
                    "group_id_hash": "g1",
                    "session_id": "g1-s000001",
                    "split": "train_candidate",
                    "start_timestamp": 1785400000000,
                    "speaker_count": 2,
                    "char_count": 20,
                    "osu_keyword_count": 0,
                    "has_media": False,
                    "media_dependent": False,
                    "privacy_risk": "low",
                    "pii_types": [],
                    "text_sanitized": "S1 hi\nS2 hello",
                },
                ensure_ascii=False,
            )
            + "\n",
            encoding="utf-8",
        )
        ann.write_text("", encoding="utf-8")
        n = build_sheet_xlsx(src, out, ann)
        self.assertEqual(n, 1)
        self.assertTrue(out.exists())
        from openpyxl import load_workbook

        wb = load_workbook(out)
        ws = wb.active
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws["A1"].value, "窗口ID")
        self.assertEqual(ws.freeze_panes, "A2")
        cell = ws["F2"].value
        self.assertIn("S2 hello", cell)
        self.assertNotIn("\x14", cell)


if __name__ == "__main__":
    unittest.main()
